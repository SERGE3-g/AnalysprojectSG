import tkinter as tk
import traceback
from tkinter import ttk, filedialog, messagebox
import threading
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook
import os
import re
import json
import queue as Queue
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


class FileTab(ttk.Frame):
    def __init__(self, parent, current_user, user_role):
        super().__init__(parent)
        self.current_user = current_user
        self.user_role = user_role

        # Variabili
        self.selected_files = []
        self.result_queue = Queue.Queue()
        self.log_queue = Queue.Queue()

        # Stile
        self.setup_style()

        # Creazione interfaccia
        self.create_gui()

        # Avvio processamento code (ascolta costantemente le code)
        self.process_queues()

    def setup_style(self):
        """Configura lo stile dell'interfaccia"""
        self.style = ttk.Style()
        self.style.configure('Header.TLabel',
                             font=("Arial", 20, "bold"),
                             padding=10)
        self.style.configure('Title.TLabel',
                             font=("Arial", 14),
                             padding=5)

    def create_gui(self):
        """Crea l'interfaccia grafica"""
        # Frame principale con divisione orizzontale
        main_frame = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Pannello sinistro (controlli)
        left_panel = self.create_left_panel(main_frame)
        main_frame.add(left_panel, weight=1)

        # Pannello destro (risultati)
        right_panel = self.create_right_panel(main_frame)
        main_frame.add(right_panel, weight=3)

    def create_left_panel(self, parent):
        """Crea il pannello sinistro con i controlli"""
        panel = ttk.Frame(parent)

        # Header
        ttk.Label(
            panel,
            text="Analizzatore File",
            style='Header.TLabel'
        ).pack(fill=tk.X)

        # Selezione file
        file_frame = ttk.LabelFrame(panel, text="File", padding=5)
        file_frame.pack(fill=tk.X, padx=5, pady=5)

        self.file_entry = ttk.Entry(file_frame)
        self.file_entry.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)

        ttk.Button(
            file_frame,
            text="Sfoglia",
            command=self.browse_files
        ).pack(side=tk.RIGHT)

        # Opzioni
        options_frame = ttk.LabelFrame(panel, text="Opzioni", padding=5)
        options_frame.pack(fill=tk.X, padx=5, pady=5)

        # Tipo analisi
        self.analysis_type = tk.StringVar(value="standard")
        ttk.Radiobutton(
            options_frame,
            text="Analisi Standard",
            variable=self.analysis_type,
            value="standard"
        ).pack(anchor=tk.W)
        ttk.Radiobutton(
            options_frame,
            text="Analisi Dettagliata",
            variable=self.analysis_type,
            value="detailed"
        ).pack(anchor=tk.W)

        # Filtri
        filters_frame = ttk.LabelFrame(options_frame, text="Filtri", padding=5)
        filters_frame.pack(fill=tk.X, pady=5)

        ttk.Label(filters_frame, text="Importo minimo:").pack(anchor=tk.W)
        self.min_amount = ttk.Entry(filters_frame)
        self.min_amount.pack(fill=tk.X, padx=5)

        ttk.Label(filters_frame, text="Importo massimo:").pack(anchor=tk.W)
        self.max_amount = ttk.Entry(filters_frame)
        self.max_amount.pack(fill=tk.X, padx=5)

        # Bottoni
        button_frame = ttk.Frame(panel)
        button_frame.pack(fill=tk.X, padx=5, pady=10)

        self.start_button = ttk.Button(
            button_frame,
            text="Avvia Analisi",
            command=self.start_analysis
        )
        self.start_button.pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="Reset",
            command=self.reset_gui
        ).pack(side=tk.LEFT)

        return panel

    def create_right_panel(self, parent):
        """Crea il pannello destro con i risultati"""
        panel = ttk.Frame(parent)

        # Progress bar
        progress_frame = ttk.Frame(panel)
        progress_frame.pack(fill=tk.X, padx=5, pady=5)

        self.progress_var = tk.DoubleVar()
        self.progress_label = ttk.Label(progress_frame, text="")
        self.progress_label.pack()

        self.progress_bar = ttk.Progressbar(
            progress_frame,
            orient=tk.HORIZONTAL,
            mode='determinate',
            variable=self.progress_var
        )
        self.progress_bar.pack(fill=tk.X)

        # Notebook per risultati
        self.results_notebook = ttk.Notebook(panel)
        self.results_notebook.pack(fill=tk.BOTH, expand=True)

        # Tab risultati
        self.create_results_tab()

        # Tab grafico
        self.create_chart_tab()

        # Tab log
        self.create_log_tab()

        return panel

    def create_results_tab(self):
        """Crea il tab dei risultati"""
        results_frame = ttk.Frame(self.results_notebook)
        self.results_notebook.add(results_frame, text="Risultati")

        # Toolbar
        toolbar = ttk.Frame(results_frame)
        toolbar.pack(fill=tk.X, pady=5)

        ttk.Button(
            toolbar,
            text="Esporta Excel",
            command=self.export_excel
        ).pack(side=tk.LEFT, padx=2)

        ttk.Button(
            toolbar,
            text="Mostra Statistiche",
            command=self.show_statistics
        ).pack(side=tk.LEFT, padx=2)

        # Tabella risultati
        columns = (
            'File', 'NbOfTxs', 'TtlIntrBkSttlmAmt',
            'Data Analisi', 'Stato'
        )
        self.results_tree = ttk.Treeview(
            results_frame,
            columns=columns,
            show='headings'
        )

        for col in columns:
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=120)

        # Scrollbar
        scrollbar = ttk.Scrollbar(
            results_frame,
            orient=tk.VERTICAL,
            command=self.results_tree.yview
        )
        self.results_tree.configure(yscrollcommand=scrollbar.set)

        self.results_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def create_chart_tab(self):
        """Crea il tab del grafico"""
        chart_frame = ttk.Frame(self.results_notebook)
        self.results_notebook.add(chart_frame, text="Grafico")

        # Tipo di grafico
        control_frame = ttk.Frame(chart_frame)
        control_frame.pack(fill=tk.X)

        self.chart_type = tk.StringVar(value="bar")
        ttk.Radiobutton(
            control_frame,
            text="Barre",
            variable=self.chart_type,
            value="bar",
            command=self.update_chart
        ).pack(side=tk.LEFT)

        ttk.Radiobutton(
            control_frame,
            text="Linee",
            variable=self.chart_type,
            value="line",
            command=self.update_chart
        ).pack(side=tk.LEFT)

        # Area grafico
        self.figure = Figure(figsize=(6, 4), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.figure, master=chart_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def create_log_tab(self):
        """Crea il tab del log"""
        log_frame = ttk.Frame(self.results_notebook)
        self.results_notebook.add(log_frame, text="Log")

        self.log_text = tk.Text(log_frame, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(
            log_frame,
            orient=tk.VERTICAL,
            command=self.log_text.yview
        )
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Tag per colorare i messaggi
        self.log_text.tag_configure("ERROR", foreground="red")
        self.log_text.tag_configure("WARNING", foreground="orange")
        self.log_text.tag_configure("SUCCESS", foreground="green")

    def process_queues(self):
        """Processa le code dei risultati e dei log, periodicamente."""
        # Processa coda risultati
        try:
            while True:
                action, data = self.result_queue.get_nowait()
                if action == "update_tree":
                    self.update_results_tree(data)
                elif action == "update_chart":
                    self.update_chart(data)
        except Queue.Empty:
            pass

        # Processa coda log
        try:
            while True:
                level, message = self.log_queue.get_nowait()
                self.add_log(level, message)
        except Queue.Empty:
            pass

        # Rischedula questo metodo
        self.after(100, self.process_queues)

    def browse_files(self):
        """Gestisce la selezione dei file"""
        files = filedialog.askopenfilenames(
            filetypes=[
                ("File supportati", "*.txt *.xml *.json"),
                ("File XML", "*.xml"),
                ("File JSON", "*.json"),
                ("File di testo", "*.txt")
            ]
        )
        if files:
            self.selected_files = list(files)
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, "; ".join(files))
            self.log_queue.put(("INFO", f"Selezionati {len(files)} file"))

    def start_analysis(self):
        """Avvia l'analisi dei file in un thread separato"""
        if not self.selected_files:
            messagebox.showwarning("Attenzione", "Seleziona almeno un file!")
            return

        # Disabilita il bottone di avvio
        self.start_button.config(state=tk.DISABLED)

        # Pulisci le visualizzazioni precedenti
        self.results_tree.delete(*self.results_tree.get_children())
        self.progress_bar['value'] = 0
        self.progress_label.config(text="")
        self.figure.clear()
        self.canvas.draw()
        self.log_text.delete(1.0, tk.END)

        # Nome file output
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_excel = f"SDD_CHECK_{now}.xlsx"

        # Avvia thread analisi
        threading.Thread(
            target=self.process_files,
            args=(self.selected_files, output_excel),
            daemon=True
        ).start()

    def process_files(self, files, output_excel):
        """Elabora i file selezionati (logica in un thread separato)"""
        try:
            self.log_queue.put(("INFO", "Avvio analisi..."))
            data = []
            total_files = len(files)
            errors = []

            for i, file_path in enumerate(files, 1):
                # Aggiorna barra progresso
                self.after(0, self.update_progress, i, total_files)

                try:
                    # Analizza file
                    results = self.parse_file(file_path)
                    results = self.apply_filters(results)

                    if results:
                        file_data = {
                            'file': os.path.basename(file_path),
                            'results': results,
                            'timestamp': datetime.now(),
                            'status': 'OK'
                        }
                        data.append(file_data)
                        # Aggiorna la tabella
                        self.after(0, lambda d=file_data:
                                   self.result_queue.put(("update_tree", d)))

                except Exception as e:
                    errors.append((file_path, str(e)))
                    self.after(0, lambda msg=f"Errore nell'analisi di {file_path}: {str(e)}:" + "\n\n" + traceback.format_exc():
                                             self.log_queue.put(("ERROR", msg)))

            if data:
                # Crea report Excel
                self.create_excel_report(data, output_excel)
                # Aggiorna grafico
                self.after(0, lambda d=data: self.result_queue.put(("update_chart", d)))
                self.after(0, lambda: self.log_queue.put(
                    ("SUCCESS", f"Analisi completata. File salvato: {output_excel}")))
                self.after(0, lambda: messagebox.showinfo(
                    "Successo", f"Analisi completata. File salvato: {output_excel}"))
            else:
                self.after(0, lambda: messagebox.showwarning(
                    "Attenzione", "Nessun dato trovato nei file analizzati"))

            if errors:
                self.after(0, lambda: self.show_errors_dialog(errors))

        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Errore", str(e)))
        finally:
            self.after(0, lambda: self.start_button.config(state=tk.NORMAL))

    def parse_file(self, file_path):
        """Analizza il file per estrarre i valori desiderati"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()

            # Regex di esempio
            txs_matches = re.findall(r'<NbOfTxs>(\d+)</NbOfTxs>', content)
            amt_matches = re.findall(r'<TtlIntrBkSttlmAmt[^>]*>([0-9.]+)</TtlIntrBkSttlmAmt>', content)

            results = []
            for txs, amt in zip(txs_matches, amt_matches):
                results.append({
                    'NbOfTxs': int(txs),
                    'TtlIntrBkSttlmAmt': float(amt)
                })
            return results

        except Exception as e:
            self.log_queue.put(("ERROR", f"Errore nel parsing del file {file_path}: {str(e)}"))
            raise

    def apply_filters(self, results):
        """Applica eventuali filtri ai risultati (importo minimo/massimo)"""
        if not results:
            return results

        filtered = list(results)  # copia

        try:
            # Filtro importo minimo
            if self.min_amount.get():
                min_val = float(self.min_amount.get())
                filtered = [r for r in filtered if r['TtlIntrBkSttlmAmt'] >= min_val]

            # Filtro importo massimo
            if self.max_amount.get():
                max_val = float(self.max_amount.get())
                filtered = [r for r in filtered if r['TtlIntrBkSttlmAmt'] <= max_val]

        except ValueError:
            self.log_queue.put(("ERROR", "Errore nei filtri: i valori devono essere numerici"))
            raise ValueError("I valori dei filtri devono essere numerici")

        return filtered

    def create_excel_report(self, data, output_file):
        """Crea il report Excel con i risultati"""
        try:
            # Prepara i dati per il DataFrame
            rows = []
            for file_data in data:
                for result in file_data['results']:
                    rows.append({
                        'File': file_data['file'],
                        'NbOfTxs': result['NbOfTxs'],
                        'TtlIntrBkSttlmAmt': result['TtlIntrBkSttlmAmt'],
                        'Data Analisi': file_data['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                        'Stato': file_data['status']
                    })

            df = pd.DataFrame(rows)

            # Aggiungi riga totali
            if not df.empty:
                totals = pd.DataFrame([{
                    'File': 'TOTALE',
                    'NbOfTxs': df['NbOfTxs'].sum(),
                    'TtlIntrBkSttlmAmt': df['TtlIntrBkSttlmAmt'].sum(),
                    'Data Analisi': '',
                    'Stato': ''
                }])
                df = pd.concat([df, totals], ignore_index=True)

            # Salva Excel
            df.to_excel(output_file, index=False)
            # Applica stili
            self.apply_excel_styles(output_file)

        except Exception as e:
            self.log_queue.put(("ERROR", f"Errore nella creazione del report Excel: {str(e)}"))
            raise

    def apply_excel_styles(self, file_path):
        """Applica stili al file Excel con openpyxl"""
        wb = load_workbook(file_path)
        ws = wb.active

        # Stili
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center")

        # Intestazione
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_align

        # Larghezza colonne
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                value_len = len(str(cell.value)) if cell.value else 0
                if value_len > max_length:
                    max_length = value_len
            ws.column_dimensions[column_letter].width = max_length + 2

        # Riga totale (ultima riga)
        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = bold_font
            cell.alignment = center_align

        wb.save(file_path)

    def update_results_tree(self, file_data):
        """Aggiorna la tabella dei risultati"""
        for result in file_data['results']:
            self.results_tree.insert('', 'end', values=(
                file_data['file'],
                result['NbOfTxs'],
                result['TtlIntrBkSttlmAmt'],
                file_data['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                file_data['status']
            ))

    def update_chart(self, data=None):
        """Aggiorna il grafico in base ai dati"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)

        if not data:
            data = self._collect_tree_data()

        if not data:
            self.canvas.draw()
            return

        # Prepara i dati per il grafico
        files = []
        amounts = []

        for file_data in data:
            for result in file_data['results']:
                files.append(file_data['file'])
                amounts.append(result['TtlIntrBkSttlmAmt'])

        if self.chart_type.get() == "bar":
            ax.bar(range(len(files)), amounts)
        else:  # line
            ax.plot(range(len(files)), amounts, marker='o')

        ax.set_xticks(range(len(files)))
        ax.set_xticklabels(files, rotation=45, ha='right')
        ax.set_title('Analisi Importi per File')
        ax.set_ylabel('Importo Totale')

        self.figure.tight_layout()
        self.canvas.draw()

    def _collect_tree_data(self):
        """Raccoglie i dati dalla tabella per il grafico (lettura diretta dalla Treeview)"""
        data = []
        for item in self.results_tree.get_children():
            values = self.results_tree.item(item)['values']
            if len(values) < 5:
                continue
            try:
                # Converte i valori
                nb_of_txs = int(values[1])
                amount = float(values[2])
                dt = datetime.strptime(values[3], '%Y-%m-%d %H:%M:%S')
                status = values[4]

                data.append({
                    'file': values[0],
                    'results': [{
                        'NbOfTxs': nb_of_txs,
                        'TtlIntrBkSttlmAmt': amount
                    }],
                    'timestamp': dt,
                    'status': status
                })
            except (ValueError, TypeError):
                continue
        return data

    def update_progress(self, current, total):
        """Aggiorna la barra di progresso"""
        progress = (current / total) * 100
        self.progress_var.set(progress)
        self.progress_label.config(text=f"Analizzato {current} di {total} file")
        self.update_idletasks()

    def add_log(self, level, message):
        """Aggiunge un messaggio al log"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.insert('end', f"[{timestamp}] {level}: {message}\n", level)
        self.log_text.see('end')

    def show_statistics(self):
        """Mostra una finestra con statistiche dettagliate"""
        data = self._collect_tree_data()
        if not data:
            messagebox.showinfo("Info", "Nessun dato disponibile per le statistiche")
            return

        stats_window = tk.Toplevel(self)
        stats_window.title("Statistiche")
        stats_window.geometry("400x300")

        try:
            # Estrai e converti i valori
            all_amounts = [float(r['TtlIntrBkSttlmAmt']) for d in data for r in d['results']]
            all_txs = [int(r['NbOfTxs']) for d in data for r in d['results']]

            if not all_amounts or not all_txs:
                messagebox.showwarning("Attenzione", "Nessun dato valido per le statistiche")
                stats_window.destroy()
                return

            stats_text = f"""
            Statistiche Analisi:

            Numero file analizzati: {len(data)}

            Importi:
            - Totale: {sum(all_amounts):,.2f}
            - Media: {sum(all_amounts) / len(all_amounts):,.2f}
            - Minimo: {min(all_amounts):,.2f}
            - Massimo: {max(all_amounts):,.2f}

            Transazioni:
            - Totale: {sum(all_txs):,d}
            - Media: {sum(all_txs) / len(all_txs):,.2f}
            - Minimo: {min(all_txs):,d}
            - Massimo: {max(all_txs):,d}
            """

            text = tk.Text(stats_window, wrap=tk.WORD, padx=10, pady=10)
            text.pack(fill=tk.BOTH, expand=True)
            text.insert('1.0', stats_text)
            text.config(state='disabled')

        except Exception as e:
            messagebox.showerror("Errore", f"Errore nel calcolo delle statistiche: {str(e)}")
            stats_window.destroy()

    def show_errors_dialog(self, errors):
        """Mostra una finestra con gli errori riscontrati"""
        dialog = tk.Toplevel(self)
        dialog.title("Errori Riscontrati")
        dialog.geometry("500x300")

        text = tk.Text(dialog, wrap=tk.WORD, padx=10, pady=10)
        text.pack(fill=tk.BOTH, expand=True)

        text.insert('1.0', "Errori durante l'analisi:\n\n")
        for file_path, error in errors:
            text.insert('end', f"File: {file_path}\nErrore: {error}\n\n")

        text.config(state='disabled')

    def reset_gui(self):
        """Resetta l'interfaccia"""
        self.file_entry.delete(0, tk.END)
        self.min_amount.delete(0, tk.END)
        self.max_amount.delete(0, tk.END)
        self.progress_bar['value'] = 0
        self.progress_label.config(text="")
        self.results_tree.delete(*self.results_tree.get_children())
        self.update_chart()
        self.selected_files = []
        self.log_queue.put(("INFO", "Interfaccia resettata"))

    def reset_results(self):
        """Resetta solo i risultati e le visualizzazioni, senza toccare i campi input"""
        self.results_tree.delete(*self.results_tree.get_children())
        self.progress_bar['value'] = 0
        self.progress_label.config(text="")
        self.figure.clear()
        self.canvas.draw()
        self.log_text.delete(1.0, tk.END)

        # Svuota eventuali code di messaggi
        while not self.result_queue.empty():
            try:
                self.result_queue.get_nowait()
            except Queue.Empty:
                pass

        while not self.log_queue.empty():
            try:
                self.log_queue.get_nowait()
            except Queue.Empty:
                pass

        self.log_queue.put(("INFO", "Risultati resettati"))

    def export_excel(self):
        """Esporta i dati correnti in Excel"""
        data = self._collect_tree_data()
        if not data:
            messagebox.showinfo("Info", "Nessun dato da esportare")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"analisi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if file_path:
            try:
                self.create_excel_report(data, file_path)
                messagebox.showinfo("Successo", f"Dati esportati in: {file_path}")
            except Exception as e:
                messagebox.showerror("Errore", f"Errore nell'esportazione: {str(e)}")

    def search(self, text):
        """Ricerca semplice nel tab (usata da una ricerca globale)"""
        results = []
        # Cerca nel log
        log_content = self.log_text.get(1.0, tk.END)
        if text.lower() in log_content.lower():
            results.append(("Log FileTab", log_content.strip()))

        # Se vuoi anche cercare nei risultati, puoi aggiungere:
        for item in self.results_tree.get_children():
            row_vals = self.results_tree.item(item)['values']
            row_str = " ".join(str(v) for v in row_vals)
            if text.lower() in row_str.lower():
                results.append(("Risultati FileTab", row_str))

        return results
